import csv
from xlrd import open_workbook
from utils import get_content
import pathlib
import codecs
import openpyxl

def csv_reader(path):
    content = []
    spamreader = codecs.decode(path.read(), 'UTF-8').split('\n')
    for row in spamreader:
        content.append(', '.join(row))
    return spamreader

def excel_reader(path):
    split_tup = pathlib.Path(path.name).suffix
    file_extension = split_tup.lower()
    if file_extension =='.xlsx':
        content = []
        workbook = openpyxl.load_workbook(path)
        for sheet_name in workbook.get_sheet_names():
            sheet = workbook.get_sheet_by_name(sheet_name)
            for _row in range(1,sheet.max_row+1):
                for _column in range(1,sheet.max_column+1):
                    call_value = sheet.cell(_row, _column).value
                    content.append(call_value)
    if file_extension =='.xls':
        content = []
        book = open_workbook(path.name, file_contents=path.read())
        for name in book.sheet_names():
            spread_sheet = book.sheet_by_name(name)
            row_count = spread_sheet.nrows
            col_count = spread_sheet.ncols
            for cur_row in range(row_count):
                for cur_col in range(col_count):
                    call_value = spread_sheet.cell(cur_row, cur_col)
                    content.append(call_value.value)
    return content

def html_reader(path):
    content = get_content.get_content_from_global_html_beta(path)
    if content:
        return content
    else:
        return get_content.get_content_from_html_sentences(path)

def pdf_reader(path):
    return get_content.get_content_from_pdf(path)

def txt_reader(path):
    return get_content.get_content_from_txt(path)

def doc_reader(path):
    return get_content.get_content_from_doc(path)

def docx_reader(path):
    return get_content.get_content_from_docx(path)

def rtf_reader(path):
    return get_content.get_content_from_rtf(path)

def generic_reader(path):
    split_tup = pathlib.Path(path.name).suffix
    file_extension = split_tup.lower()
    methods = {'.csv':csv_reader, '.html':html_reader, '.htm':html_reader, '.pdf':pdf_reader, '.txt':txt_reader, '.xls':excel_reader, '.xlsx':excel_reader,
                '.doc':doc_reader, '.docx':docx_reader, '.rtf':rtf_reader}
    return methods[file_extension](path)



if __name__ == "__main__":
    path = '/FE_documents/ISIN/US3136G02L40.pdf'
    print(generic_reader(path))