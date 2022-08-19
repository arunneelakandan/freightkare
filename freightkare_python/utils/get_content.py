from __future__ import division, unicode_literals 
import PyPDF2
from bs4 import BeautifulSoup
import codecs
import re
from utils import global_variables
from utils import base_class
import os
import openpyxl
import csv
from xlrd import open_workbook
import json  
from statistics import mode
import textract
import docx2txt
import io

def clean_dict_snake_case(dict):
    return {
        base_class.snake_case(base_class.remove_spl_char2(key)): value
        for key, value in dict.items()
    }

def extract_from_mails(html):
    Extracted_all_row = []
    Extracted_dict = {}
    html_tags = BeautifulSoup(html, 'html5lib')
    html_tags = html_tags.find('div', attrs = {'class':'gmail_quote'})
    try:
        pre_content_paragraph = html_tags.findAll('div')[1]
        content_paragraph = pre_content_paragraph.findAll('div')[0]

        if not content_paragraph.find('table'):
            print('div')
            for i in range(1,len(content_paragraph.findAll('div'))):
                extracted_rows = content_paragraph.findAll('div')[i].text
                extracted_rows = extracted_rows.split(':',1)
                Extracted_all_row.append(extracted_rows)

            for j in range(len(Extracted_all_row)):
                try:
                    if '   ' in Extracted_all_row[j][0]:
                        Extracted_all_row[j-1][1] = Extracted_all_row[j-1][1] + Extracted_all_row[j][0].strip()
                except:
                    pass  

            for i in Extracted_all_row:
                if len(i) > 1:
                    Extracted_dict[i[0]] = i[1].strip()
        else:
            table_content = pre_content_paragraph.find('tbody')

            for rows in table_content.findAll('tr'):
                if len(rows.findAll('td')) > 1:
                    table_row_col1 = rows.findAll('td')[0].text.replace(':','')
                    table_row_col2 = rows.findAll('td')[1].text
                    Extracted_dict[table_row_col1] = table_row_col2
        Extracted_dict['factentry http status'] = 200
    except: 
        Extracted_dict['factentry http status'] = 100
        return Extracted_dict
    return clean_dict_snake_case(Extracted_dict)

def get_content_from_html(path_file_name):
    content = []
    try:
        page = codecs.open(path_file_name, 'r')
        soup = BeautifulSoup(page, 'html.parser')
        content_p = soup.find_all(['p'])
        for p in content_p: 
            content.append(p.getText())
        content = [c.replace('\u00a0', ' ') for c in content]
        content = [c.replace('\r', ' ') for c in content]
        content = [c.replace('\n', ' ') for c in content]
        while("" in content):
            content.remove("")
        while(" " in content):
            content.remove(" ")
        while("-" in content):
            content.remove("-")
    except (FileNotFoundError, UnicodeDecodeError):
        pass
    return content


def get_content_from_html_beta(path_file_name):
    print(path_file_name)
    content = []
    try:
        with codecs.open(path_file_name, 'r', encoding='utf-8', errors='ignore') as f:
        # with open(path_file_name, "r") as f:
            content = get_content_from_html_file(f)
    except (FileNotFoundError, UnicodeDecodeError):
        pass
    return content

def get_content_from_global_html_beta(path_file_name):
    content = []
    # with codecs.open(path_file_name, 'r', encoding='utf-8', errors='ignore') as f:
        # with open(path_file_name, "r") as f:
    content = get_content_from_html_file(path_file_name)
    return content

def get_content_from_html_beta_bold(path_file_name):
    content = []
    try:
        with open(path_file_name, "r") as f:
            soup = BeautifulSoup(f, 'html.parser')
        content_b = soup.find_all(['b'])
        for b in content_b:
            content.append(b.getText())
        content = [c.replace("\u00a0", " ") for c in content]
        content = [c.replace("\r", " ") for c in content]
        content = [c.replace("\n", " ") for c in content]
        content = [c.replace("\xa0", " ") for c in content]
        content = [re.sub(' +', ' ', c) for c in content]
        # while("" in content) :
        #     content.remove("")
        while("&nbsp;" in content):
            content.remove("&nbsp;")
        while(" " in content):
            content.remove(" ")
        while("-" in content):
            content.remove("-")
    except (FileNotFoundError, UnicodeDecodeError):
        pass
    return content


def get_content_from_pdf(path_file_name):
    content = []
    # if (
    #     'Complete document is searchable'
    #     in is_scanned.get_pdf_searchable_pages(path_file_name)['comments']
    # ):
    # with open(path_file_name, 'rb') as pdfFileObj:
    pdfReader = PyPDF2.PdfFileReader(io.BytesIO(path_file_name.read()))
    for i in range(pdfReader.numPages):
        # for i in range(0,2):
        # print ('************ Page No - '+str(i)+' ************')
        # creating a page object
        pageObj = pdfReader.getPage(i)
        # extracting text from page
        lines = pageObj.extractText().split('\n')
        # while("" in lines) :
        #     lines.remove("")
        # while(" " in lines):
        #     lines.remove(" ")
        # data_string = str(data_string)
        for i in range(len(lines)):
            content.append(lines[i])
                # print (lines[i])
    while(" " in content):
        content.remove(" ")
    return content


def get_content_from_txt(path_file_name):
    content = []
    # with open(path_file_name.read(), "r") as fp:
    content = codecs.decode(path_file_name.read(), 'UTF-8').split('\n')
    content = [c.replace('\xa0', ' ') for c in content]
    content = [re.sub(' +', ' ', c) for c in content]
    return content


def get_content_from_html_file(file):
    content = []
    try:
        soup = BeautifulSoup(file, 'html.parser')
        body = soup.body
        content = str(body)
        content = striphtml(content)
        content = content.replace('\xa0',' ')
        lines = content.split('\n')
        
        while('' in lines):
            lines.remove('')
        while(" " in lines):
            lines.remove(" ")
        while("-" in lines):
            lines.remove("-")
        while("." in lines):
            lines.remove(".")
        while('§' in lines):
            lines.remove('§')
        while(':' in lines):
            lines.remove(':')
            
        lines = [c.replace('\u00a0', ' ') for c in lines]
        lines = [c.replace('\r', ' ') for c in lines]
        lines = [c.replace('\n', ' ') for c in lines]
        lines = [c.replace('\xa0', ' ') for c in lines]
        lines = [re.sub(' +', ' ', c) for c in lines]
        lines = [c.replace('"', '') for c in lines]
        lines = [c.replace('„', '') for c in lines]
        # lines = [base_class.remove_non_ascii(c) for c in lines]
        # while("" in lines) :
        #     lines.remove("")
        return lines
        
    except :
        pass
    
    # lines = ','.join(lines)
    


def get_content_from_global_html(file):
    content = []
    soup = BeautifulSoup(file, 'html.parser')
    content_td = soup.find_all(['td'])
    # content_th = soup.find_all(['th'])
    content_li = soup.find_all(['li'])
    content_span = soup.find_all(['span'])
    content_p = soup.find_all(['p'])
    content_div = soup.find_all(['div'])
    content = [td.getText() for td in content_td]
    for span in content_span:
        content.append(span.getText())
    for p in content_p:
        content.append(p.getText())
    for li in content_li:
        content.append(li.getText())
    for div in content_div:
        content.append(div.getText())
    content = [c.replace('\u00a0', '') for c in content]
    content = [c.replace('\r', '') for c in content]
    content = [c.replace('\n', '') for c in content]
    content = [c.replace('\xa0', ' ') for c in content]
    content = [re.sub(' +', ' ', c) for c in content]
    while(" " in content):
        content.remove(" ")
    while("-" in content):
        content.remove("-")
    while("" in content):
        content.remove("")
    return content

def get_content_from_pdf_file(file):
    # pdfReader = PyPDF2.PdfFileReader(file.read().decode("latin-1"))
    
    content = []

    with open(file, 'rb') as pdfFileObj:
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        for i in range(pdfReader.numPages):
            # for i in range(0,2):
            # print ('************ Page No - '+str(i)+' ************')
            # creating a page object
            pageObj = pdfReader.getPage(i)
            # extracting text from page
            lines = pageObj.extractText().split('\n')
            # while("" in lines) :
            #     lines.remove("")
            # while(" " in lines):
            #     lines.remove(" ")
            # data_string = str(data_string)
            for i in range(len(lines)):
                content.append(lines[i])
                # print (lines[i])
        while(" " in content):
            content.remove(" ")
    
    return content

def get_content_from_csv_file(file):  # sourcery no-metrics
    synonyms = global_variables.synonyms
    synonyms['cusip']=['cusip']
        
    all_sheets_extracted_dict = []
    extracted_dict = {}
    
    sheet = csv.reader(file.read())
    sheet = list(sheet) 
    header_found = False
    header_row, header_index = [],''
    keys = list(synonyms.keys())
    values = [j for i in synonyms.values() for j in i]
    key_values = keys + values
    
    #For Finding Header in list of rows present in the sheet with help of otto synonyms
    
    for rows in sheet:
        for synonym_key, synonym_value in synonyms.items():
            if header_found == False:
                check_header_1 = [i for i in rows if i.lower() in synonym_value]
                check_header_2 = any(synonym_key in s.lower() for s in rows)
                if len(check_header_1) >= 1 or check_header_2 == True:
                    header_row = rows
                    header_index = sheet.index(rows)
                    header_found = True
                    break
    
    
    if len(header_row) >= 1:
        
        extracted_dict['dataset'] = []
        extracted_dict['exception']=[]  
        for rows in range(int(header_index),len(sheet)):
            exception = {}
            data = {}
            if len(sheet[rows]) == len(header_row):
                for num in range(len(sheet[rows])):
                    if len([syn_vals for syn_vals in key_values if syn_vals in header_row[num].lower()]) >= 1:
                        data[[syn_vals for syn_vals in key_values if syn_vals in header_row[num].lower()][0]] = sheet[rows][num] 
                    else:
                        exception[header_row[num].lower()] = sheet[rows][num]
            extracted_dict['dataset'].append(data)
        extracted_dict['exception'].append(exception)
    
    else:
        header_row, header_index = [],''
        mode_row = mode([len(rows) for rows in sheet])
        for rows in sheet:   
            if len(rows) == mode_row:
                header_row = rows
                header_index = sheet.index(rows)
                break
            
        extracted_dict['dataset'] = []
        extracted_dict['exception']=[]  
        for rows in range(int(header_index),len(sheet)):
            exception = {}
            data = {}
            if len(sheet[rows]) == len(header_row):
                for num in range(len(sheet[rows])):
                    exception[header_row[num].lower()] = sheet[rows+5][num]
                break
        extracted_dict['exception'].append(exception)        

    all_sheets_extracted_dict.append(extracted_dict)
             
    return all_sheets_extracted_dict    

def get_content_from_xls_file(file):
    synonyms = global_variables.synonyms
    synonyms['cusip']=['cusip']


    all_sheets_extracted_dict = []
    
    
    
    book = open_workbook(file.name, file_contents=file.read())
    for name in book.sheet_names():
        spread_sheet = book.sheet_by_name(name)
        row_count = spread_sheet.nrows
        col_count = spread_sheet.ncols
        extracted_dict = {}
        sheet=[]
        for cur_row in range(0, row_count):
            list_of_value=[]
            for cur_col in range(0, col_count):
                cell = spread_sheet.cell(cur_row, cur_col)
                list_of_value.append(str(cell.value).strip())
            sheet.append(list_of_value)
        
        if len(sheet) >= 1:
            header_found = False
            header_row, header_index = [],''
            keys = list(synonyms.keys())
            values = [j for i in synonyms.values() for j in i]
            key_values = keys + values

            #For Finding Header in list of rows present in the sheet with help of otto synonyms

            for rows in sheet:
                for synonym_key, synonym_value in synonyms.items():
                    if header_found == False:
                        check_header_1 = [i for i in rows if i.lower() in synonym_value]
                        check_header_2 = any(synonym_key in s.lower() for s in rows)
                        if len(check_header_1) >= 1 or check_header_2 == True:
                            header_row = rows
                            header_index = sheet.index(rows)
                            header_found = True
                            break


            if len(header_row) >= 1:
                
                extracted_dict['dataset'] = []
                extracted_dict['exception']=[]  
                for rows in range(int(header_index),len(sheet)):
                    exception = {}
                    data = {}
                    if len(sheet[rows]) == len(header_row):
                        for num in range(len(sheet[rows])):
                            if len([syn_vals for syn_vals in key_values if syn_vals in header_row[num].lower()]) >= 1:
                                data[[syn_vals for syn_vals in key_values if syn_vals in header_row[num].lower()][0]] = sheet[rows][num] 
                            else:
                                exception[header_row[num].lower()] = sheet[rows][num]
                    extracted_dict['dataset'].append(data)
                extracted_dict['exception'].append(exception)

            else:
                header_row, header_index = [],''
                mode_row = mode([len(rows) for rows in sheet])
                for rows in sheet:   
                    if len(rows) == mode_row:
                        header_row = rows
                        header_index = sheet.index(rows)
                        break
                    
                extracted_dict['dataset'] = []
                extracted_dict['exception']=[]  
                for rows in range(int(header_index),len(sheet)):
                    exception = {}
                    data = {}
                    if len(sheet[rows]) == len(header_row):
                        for num in range(len(sheet[rows])):
                            exception[header_row[num].lower()] = sheet[rows+5][num]
                        break
                extracted_dict['exception'].append(exception)                        
            
            all_sheets_extracted_dict.append(extracted_dict)
        
        else:
            pass    
    
    return all_sheets_extracted_dict
    
    
def get_content_from_xlsx_file(file):
    synonyms = global_variables.synonyms
    synonyms['cusip']=['cusip']


    all_sheets_extracted_dict = []
    
    
    
    # load excel with its path
    wrkbk = openpyxl.load_workbook(file)
    
    for spread_sheet in wrkbk.worksheets:
        extracted_dict = {}
        sheet = []  
        # iterate through excel and display data
        for row in spread_sheet.iter_rows(min_row=1, min_col=1):
            sheet_row = []
            for cell in row:
                if cell.value != None :
                    sheet_row.append(str(cell.value))
                else:
                    sheet_row.append('') 
            sheet.append(sheet_row)
        
        if len(sheet) >= 1:
            header_found = False
            header_row, header_index = [],''
            keys = list(synonyms.keys())
            values = [j for i in synonyms.values() for j in i]
            key_values = keys + values

            #For Finding Header in list of rows present in the sheet with help of otto synonyms

            for rows in sheet:
                for synonym_key, synonym_value in synonyms.items():
                    if header_found == False:
                        check_header_1 = [i for i in rows if i.lower() in synonym_value]
                        check_header_2 = any(synonym_key in s.lower() for s in rows)
                        if len(check_header_1) >= 1 or check_header_2 == True:
                            header_row = rows
                            header_index = sheet.index(rows)
                            header_found = True
                            break


            if len(header_row) >= 1:
                
                extracted_dict['dataset'] = []
                extracted_dict['exception']=[]  
                for rows in range(int(header_index),len(sheet)):
                    exception = {}
                    data = {}
                    if len(sheet[rows]) == len(header_row):
                        for num in range(len(sheet[rows])):
                            if len([syn_vals for syn_vals in key_values if syn_vals in header_row[num].lower()]) >= 1:
                                data[[syn_vals for syn_vals in key_values if syn_vals in header_row[num].lower()][0]] = sheet[rows][num] 
                            else:
                                exception[header_row[num].lower()] = sheet[rows][num]
                    extracted_dict['dataset'].append(data)
                extracted_dict['exception'].append(exception)

            else:
                header_row, header_index = [],''
                mode_row = mode([len(rows) for rows in sheet])
                for rows in sheet:   
                    if len(rows) == mode_row:
                        header_row = rows
                        header_index = sheet.index(rows)
                        break
                    
                extracted_dict['dataset'] = []
                extracted_dict['exception']=[]  
                for rows in range(int(header_index),len(sheet)):
                    exception = {}
                    data = {}
                    if len(sheet[rows]) == len(header_row):
                        for num in range(len(sheet[rows])):
                            exception[header_row[num].lower()] = sheet[rows+5][num]
                        break
                extracted_dict['exception'].append(exception)                        
            
            all_sheets_extracted_dict.append(extracted_dict)
        
        else:
            pass
            
     
    return all_sheets_extracted_dict

def striphtml(data):
    try:
        p = re.compile(r'<.*?>')
        return p.sub('', data)
    except:
        pass
    
def get_content_from_html_sentences(file):
    
    soup = codecs.open(file, "r", "utf-8")
    content = soup.read()
    content = striphtml(content)
    content = content.replace('\n','')
    content = content.replace('\t',' ')
    from nltk.tokenize import sent_tokenize

    return sent_tokenize(content.strip())

def get_content_from_text_file(file):
    
    txt_content = open(file)
    content = []
    for i in txt_content.readlines():
        j=i.replace('\n','')
        k=j.replace('\t',' ')
        L=k.strip()
        L=striphtml(L)
        content.append(L)
    
    return content

def get_content_from_html_sentences_beta(file):
    
    soup = codecs.open(file, "r", "utf-8")
    content = BeautifulSoup(soup, 'html.parser')
    content = str(content)
    content = striphtml(content)
    content = content.replace('\n',' ')
    content = content.replace('\t',' ')
    from nltk.tokenize import sent_tokenize
    content = content.replace('\xa0',' ')
    content = sent_tokenize(content.strip())
    
    return content
def get_html_content(file):
    lines = []
    try:
        soup = codecs.open(file, "r", "utf-8")
        content = BeautifulSoup(soup, 'html.parser')
        content = str(content)
        content = striphtml(content)
        content = content.replace('\xa0',' ')
        lines = content.split('\n')
        
        while('' in lines):
            lines.remove('')
        while(" " in lines):
            lines.remove(" ")
        while("-" in lines):
            lines.remove("-")
        while("." in lines):
            lines.remove(".")
        while('§' in lines):
            lines.remove('§')
        while(':' in lines):
            lines.remove(':')
            
        lines = [c.replace('\u00a0', ' ') for c in lines]
        lines = [c.replace('\r', ' ') for c in lines]
        lines = [c.replace('\n', ' ') for c in lines]
        lines = [c.replace('\xa0', ' ') for c in lines]
        lines = [re.sub(' +', ' ', c) for c in lines]
        lines = [c.replace('"', '') for c in lines]
        lines = [c.replace('„', '') for c in lines]
        # lines = [base_class.remove_non_ascii(c) for c in lines]
        # while("" in lines) :
        #     lines.remove("")
        
    except :
        pass
    
    return lines

def get_content_from_doc(file):
    result = textract.process(file.read())
    result = str(result).split('\\n')
    result = [i.strip() for i in result if i != '']
    result = [i.replace('     ',' ') for i in result]
    
    return result

def get_content_from_docx(file):
    result = docx2txt.process(file)
    result = result.split('\n')
    result = [i.strip() for i in result if i != '']
    result = [i.replace('\t',' ') for i in result]  

    return result  

def get_content_from_rtf(file):
    result = textract.process(file)
    result = str(result).split('\\n')
    result = [i.strip() for i in result]
    result = [i.strip() for i in result if i != '']

    return result 